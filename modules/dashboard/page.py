"""날짜별 대시보드 — 자재 상·하차 반입/반출 현황 테이블."""
import io
from datetime import date, timedelta

import streamlit as st
from shared.timing import measure
from supabase import Client
from config import KIND_IN, KIND_OUT
from db.models import settings_get
from shared.helpers import today_kst


_DASH_CSS = """
<style>
/* 날짜 네비 */
.st-key-dash_nav { margin-bottom: 16px !important; }
.st-key-dash_nav .stHorizontalBlock {
  gap: 4px !important;
  align-items: center !important;
  flex-wrap: nowrap !important;
}
/* 버튼 컬럼(◀◀ ◀ ▶ ▶▶): 고정 최소폭 */
.st-key-dash_nav .stHorizontalBlock > div:nth-child(1),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(2),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(4),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(5) {
  flex: 0 0 40px !important;
  min-width: 40px !important;
  max-width: 40px !important;
}
/* 6번째 컬럼(오늘 버튼): 살짝 넓게 — 한글 두 글자 표시 */
.st-key-dash_nav .stHorizontalBlock > div:nth-child(6) {
  flex: 0 0 56px !important;
  min-width: 56px !important;
  max-width: 56px !important;
}
/* 날짜 박스 컬럼: 남은 공간 차지 */
.st-key-dash_nav .stHorizontalBlock > div:nth-child(3) {
  flex: 1 1 auto !important;
  min-width: 0 !important;
}
.st-key-dash_nav button {
  height: 38px !important;
  min-height: 38px !important;
  max-height: 38px !important;
  padding: 0 !important;
  font-size: 13px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
}
.st-key-dash_nav button p,
.st-key-dash_nav button span {
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  margin: 0 !important;
  line-height: 1 !important;
}
.st-key-dash_nav [data-baseweb="input"] {
  height: 38px !important;
  min-height: 38px !important;
}
.st-key-dash_nav [data-baseweb="input"] input {
  height: 38px !important;
  line-height: 38px !important;
  padding-top: 0 !important;
  padding-bottom: 0 !important;
  text-align: center !important;
}

/* 대시보드 래퍼 */
div.dash-wrap {
  width: 100% !important;
  overflow-x: scroll !important;
  -webkit-overflow-scrolling: touch !important;
  cursor: grab !important;
}
div.dash-wrap:active { cursor: grabbing !important; }

/* 가로 스크롤바를 굵고 잡기 쉽게 */
div.dash-wrap::-webkit-scrollbar { height: 14px !important; }
div.dash-wrap::-webkit-scrollbar-track { background: #cbd5e1 !important; border-radius: 7px !important; }
div.dash-wrap::-webkit-scrollbar-thumb { background: #64748b !important; border-radius: 7px !important; border: 2px solid #cbd5e1 !important; }
div.dash-wrap::-webkit-scrollbar-thumb:hover { background: #334155 !important; }

/* 타이틀 박스 */
.dash-title-box {
  border: 2px solid #1e3a8a;
  border-radius: 4px;
  text-align: center;
  padding: 16px 8px 14px 8px;
  margin-bottom: 12px;
  width: 100%;
  box-sizing: border-box;
}
.dash-title-box h2 {
  font-size: clamp(20px, 4vw, 32px);
  font-weight: 900;
  color: #0f172a;
  margin: 0;
  letter-spacing: -0.5px;
}

/* 사이트·날짜 헤더 */
.dash-meta {
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 10px;
  font-size: 13px;
  color: #1e3a8a;
  font-weight: 600;
}

/* 테이블 */
.dash-table {
  width: 100%;
  border-collapse: collapse;
  font-size: clamp(11px, 1.8vw, 13px);
  min-width: 860px;
}
.dash-table th {
  background: #1e3a8a;
  color: #ffffff;
  padding: 8px 6px;
  text-align: center;
  border: 1px solid #1e40af;
  font-weight: 700;
  white-space: nowrap;
  vertical-align: middle;
}
.dash-table th.th-sub {
  background: #2563eb;
  font-size: 11px;
  padding: 5px 4px;
}
.dash-table td {
  padding: 7px 6px;
  border: 1px solid #cbd5e1;
  text-align: center !important;
  vertical-align: middle !important;
  color: #0f172a;
  word-break: keep-all;
  cursor: help;
}
.dash-table td:hover {
  background: #dbeafe !important;
}
@media (max-width: 480px) {
  .dash-table td {
    padding: 8px 4px;
    font-size: 10px;
    word-break: break-word;
    white-space: normal;
    vertical-align: top;
    max-width: 70px;
  }
}
.tooltip-content {
  position: fixed;
  background: #0f172a;
  color: #ffffff;
  padding: 12px 16px;
  border-radius: 6px;
  font-size: 13px;
  max-width: 280px;
  word-wrap: break-word;
  z-index: 9999;
  box-shadow: 0 4px 12px rgba(0,0,0,0.3);
  display: none;
  cursor: pointer;
}
.tooltip-content.active {
  display: block;
}
.dash-table tr:nth-child(even) td { background: #f8fafc; }
.dash-table tr:nth-child(odd)  td { background: #ffffff; }
.dash-table tr:hover td { background: #eff6ff !important; }

/* 반입/반출 배지 */
.kind-in  { color: #1d4ed8; font-weight: 700; }
.kind-out { color: #b91c1c; font-weight: 700; }

/* 합계 행 */
.dash-table tr.total-row td {
  background: #e0e7ff !important;
  font-weight: 700;
  color: #1e3a8a;
}

/* 다운로드 버튼 */
[data-testid="stDownloadButton"] button {
  white-space: nowrap !important;
  background-color: #2563eb !important;
  border-color: #2563eb !important;
  color: #ffffff !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
}
[data-testid="stDownloadButton"] button:hover {
  background-color: #1d4ed8 !important;
  border-color: #1d4ed8 !important;
}
[data-testid="stDownloadButton"] button p,
[data-testid="stDownloadButton"] button span {
  white-space: nowrap !important;
  color: #ffffff !important;
  margin: 0 !important;
  line-height: 1 !important;
}

/* 모바일: 모바일 전용 영역 숨김, 테이블 표시 (가로 스크롤 가능) */
.dash-mobile-only { display: none; }
</style>
"""


def _build_excel(reqs: list, site_name: str, date_label: str, terminal_zones: set = None) -> bytes:
    """요청 목록을 엑셀 파일로 변환하여 bytes 반환."""
    if terminal_zones is None:
        terminal_zones = set()
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "반입반출현황"

    # ── 스타일 정의 ───────────────────────────────────────────────────────
    thin = Side(style="thin", color="808080")
    thick = Side(style="medium", color="1E3A8A")
    border_all  = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    hdr_fill   = PatternFill("solid", fgColor="1E3A8A")
    sub_fill   = PatternFill("solid", fgColor="2563EB")
    total_fill = PatternFill("solid", fgColor="E0E7FF")
    even_fill  = PatternFill("solid", fgColor="F8FAFC")

    hdr_font   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    body_font  = Font(name="맑은 고딕", size=10)
    total_font = Font(name="맑은 고딕", bold=True, color="1E3A8A", size=10)
    title_font = Font(name="맑은 고딕", bold=True, size=16)
    meta_font  = Font(name="맑은 고딕", bold=True, size=10, color="1E3A8A")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── 행 1: 제목 ────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"].value = "자재 상·하차 반입/반출 현황"
    ws["A1"].font  = title_font
    ws["A1"].alignment = center
    ws["A1"].border = border_thick
    for _c in range(2, 13):  # B1:L1 — merged cell border fix
        ws.cell(row=1, column=_c).border = border_thick
    ws.row_dimensions[1].height = 44

    # ── 행 2: 현장명 / 날짜 ───────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"□ {site_name}"
    ws["A2"].font  = meta_font
    ws["A2"].alignment = left
    ws.merge_cells("G2:L2")
    ws["G2"].value = date_label
    ws["G2"].font  = meta_font
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── 행 3: 헤더 (단일행) ───────────────────────────────────────────────
    for col_letter, label in [
        ("A","No"), ("B","Zone"), ("C","터미널"), ("D","업체명"), ("E","자재종류"),
        ("F","수량"), ("G","반입·반출 차량"), ("H","상·하차 방식"),
        ("I","시간"), ("J","작업지휘자"), ("K","유도원"), ("L","담당자"),
    ]:
        ws.merge_cells(f"{col_letter}3:{col_letter}4")
        cell = ws[f"{col_letter}3"]
        cell.value = label
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = center
        cell.border = border_all
        # 병합 하단 셀(4행)에도 테두리 적용 — 세로선 누락 방지
        ws[f"{col_letter}4"].border = border_all

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    # ── 데이터 행 ─────────────────────────────────────────────────────────
    dow_map = {0:"월요일", 1:"화요일", 2:"수요일", 3:"목요일",
               4:"금요일", 5:"토요일", 6:"일요일"}
    total_cnt = 0
    for i, r in enumerate(reqs, 1):
        row_num = i + 4
        kind    = r.get("kind", KIND_IN)
        kind_lbl = "반입" if kind == KIND_IN else "반출"
        company  = r.get("company_name", "")
        item     = r.get("item_name", "")
        vcnt_raw = r.get("vehicle_count", "")
        vcnt     = f"{vcnt_raw}대" if vcnt_raw else ""
        vton     = r.get("vehicle_ton", "")
        loading  = r.get("loading_method", "")
        gate_raw   = r.get("gate", "")
        gate_parts = gate_raw.split("|", 1) if "|" in gate_raw else [gate_raw, ""]
        gate_zone  = gate_parts[0].strip()
        gate_place = gate_parts[1].strip() if len(gate_parts) > 1 else ""
        t_from = r.get("time_from", "")
        sup    = r.get("worker_supervisor", "")
        guide  = r.get("worker_guide", "")
        mgr    = r.get("worker_manager", "")
        vcnt_int = int(vcnt_raw) if str(vcnt_raw).isdigit() else 0
        total_cnt += vcnt_int

        zone_xl     = r.get("booking_zone", "")
        gate_xl_raw = r.get("gate", "")
        gate_xl_raw = "" if gate_xl_raw == "선택" else gate_xl_raw
        gate_xl     = gate_xl_raw if zone_xl in terminal_zones else "N/A"
        fill = even_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [i, zone_xl, gate_xl, company, item, vcnt, f"{kind_lbl} / {vton}", loading,
                  t_from, sup, guide, mgr]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font   = body_font
            cell.fill   = fill
            cell.border = border_all
            cell.alignment = center
        ws.row_dimensions[row_num].height = 18

    # ── 합계 행 ───────────────────────────────────────────────────────────
    total_row = len(reqs) + 5
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"].value = "합 계"
    ws[f"A{total_row}"].font  = total_font
    ws[f"A{total_row}"].fill  = total_fill
    ws[f"A{total_row}"].alignment = center
    ws[f"A{total_row}"].border = border_all
    # 병합 셀(B~E) 각각에도 테두리 적용
    for col_letter in ["B", "C", "D", "E"]:
        ws[f"{col_letter}{total_row}"].border = border_all
    ws[f"F{total_row}"].value = f"{total_cnt}대"
    ws[f"F{total_row}"].font  = total_font
    ws[f"F{total_row}"].fill  = total_fill
    ws[f"F{total_row}"].alignment = center
    ws[f"F{total_row}"].border = border_all
    for col_idx in range(7, 13):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill   = total_fill
        cell.border = border_all
    ws.row_dimensions[total_row].height = 18

    # ── 열 너비 ───────────────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter
    col_widths = [5, 8, 14, 14, 35, 10, 12, 14, 8, 12, 10, 10]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── bytes 반환 ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_report_html(reqs: list, site_name: str, date_from: str, date_to: str) -> str:
    """6/15~ 사용현황 HTML 보고서 생성."""
    from collections import defaultdict, Counter

    total      = len(reqs)
    in_reqs    = [r for r in reqs if r.get("kind") == KIND_IN]
    out_reqs   = [r for r in reqs if r.get("kind") != KIND_IN]
    done_reqs  = [r for r in reqs if r.get("status") == "DONE"]

    # 일별 집계
    daily_in:  dict = defaultdict(int)
    daily_out: dict = defaultdict(int)
    for r in reqs:
        d = (r.get("date") or "")[:10]
        if r.get("kind") == KIND_IN:
            daily_in[d]  += 1
        else:
            daily_out[d] += 1
    all_dates = sorted(set(list(daily_in.keys()) + list(daily_out.keys())))

    # 존별 집계
    zone_cnt: Counter = Counter(r.get("booking_zone") or "미지정" for r in reqs)

    # 협력사별 집계 (상위 15)
    co_cnt: Counter = Counter(r.get("company_name") or "미지정" for r in reqs)
    top_cos = co_cnt.most_common(15)

    # 시간대별 집계 (1시간 단위)
    hour_cnt: Counter = Counter()
    for r in reqs:
        tf = r.get("time_from") or ""
        if tf:
            try:
                hour_cnt[int(tf[:2])] += 1
            except Exception:
                pass

    # ── 일별 바 차트 ──────────────────────────────────────────────────────
    max_daily = max((daily_in[d] + daily_out[d]) for d in all_dates) if all_dates else 1
    BAR_H = 120
    bar_rows = ""
    for d in all_dates:
        iv = daily_in.get(d, 0)
        ov = daily_out.get(d, 0)
        tot = iv + ov
        label = d[5:]  # MM-DD
        ih = int(iv / max_daily * BAR_H) if max_daily else 0
        oh = int(ov / max_daily * BAR_H) if max_daily else 0
        bar_rows += (
            f"<div style='display:flex;flex-direction:column;align-items:center;gap:2px;min-width:34px;'>"
            f"<div style='font-size:9px;color:#64748b;'>{tot}</div>"
            f"<div style='display:flex;flex-direction:column;justify-content:flex-end;height:{BAR_H}px;gap:1px;'>"
            f"<div style='width:22px;height:{ih}px;background:#3b82f6;border-radius:2px 2px 0 0;' title='반입 {iv}'></div>"
            f"<div style='width:22px;height:{oh}px;background:#f97316;border-radius:2px 2px 0 0;' title='반출 {ov}'></div>"
            f"</div>"
            f"<div style='font-size:9px;color:#475569;'>{label}</div>"
            f"</div>"
        )

    # ── 존별 바 ───────────────────────────────────────────────────────────
    zone_colors = ["#6366f1","#0ea5e9","#10b981","#f59e0b","#ef4444","#8b5cf6","#ec4899"]
    zone_rows = ""
    max_zone = max(zone_cnt.values()) if zone_cnt else 1
    for i, (z, cnt) in enumerate(sorted(zone_cnt.items())):
        pct = int(cnt / total * 100) if total else 0
        bw  = int(cnt / max_zone * 200)
        clr = zone_colors[i % len(zone_colors)]
        zone_rows += (
            f"<tr>"
            f"<td style='padding:5px 8px;font-size:12px;font-weight:600;color:#334155;width:80px;'>{z}</td>"
            f"<td style='padding:5px 4px;'>"
            f"<div style='width:{bw}px;height:16px;background:{clr};border-radius:3px;'></div></td>"
            f"<td style='padding:5px 8px;font-size:12px;color:#64748b;'>{cnt}건 ({pct}%)</td>"
            f"</tr>"
        )

    # ── 협력사별 ──────────────────────────────────────────────────────────
    max_co = top_cos[0][1] if top_cos else 1
    co_rows = ""
    for i, (co, cnt) in enumerate(top_cos):
        pct = int(cnt / total * 100) if total else 0
        bw  = int(cnt / max_co * 200)
        clr = zone_colors[i % len(zone_colors)]
        co_rows += (
            f"<tr>"
            f"<td style='padding:5px 8px;font-size:11px;color:#334155;max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'>{co}</td>"
            f"<td style='padding:5px 4px;'>"
            f"<div style='width:{bw}px;height:14px;background:{clr};border-radius:3px;'></div></td>"
            f"<td style='padding:5px 8px;font-size:11px;color:#64748b;'>{cnt}건 ({pct}%)</td>"
            f"</tr>"
        )

    # ── 시간대별 ──────────────────────────────────────────────────────────
    max_hour = max(hour_cnt.values()) if hour_cnt else 1
    hour_rows = ""
    for h in range(5, 21):
        cnt = hour_cnt.get(h, 0)
        bw  = int(cnt / max_hour * 180) if max_hour else 0
        hour_rows += (
            f"<tr>"
            f"<td style='padding:3px 8px;font-size:11px;color:#475569;width:50px;'>{h:02d}시</td>"
            f"<td style='padding:3px 4px;'>"
            f"<div style='width:{bw}px;height:12px;background:#0ea5e9;border-radius:2px;'></div></td>"
            f"<td style='padding:3px 8px;font-size:11px;color:#64748b;'>{cnt}건</td>"
            f"</tr>"
        )

    # ── 상세 목록 ─────────────────────────────────────────────────────────
    STATUS_KO = {
        "PENDING_APPROVAL": "확정대기",
        "APPROVED": "확정완료",
        "REJECTED": "반려",
        "EXECUTING": "실행중",
        "DONE": "등록완료",
    }
    detail_rows = ""
    for r in sorted(reqs, key=lambda x: (x.get("date",""), x.get("booking_zone",""), x.get("time_from",""))):
        kind_lbl = "반입" if r.get("kind") == KIND_IN else "반출"
        kind_clr = "#3b82f6" if r.get("kind") == KIND_IN else "#f97316"
        status   = STATUS_KO.get(r.get("status",""), r.get("status",""))
        detail_rows += (
            f"<tr>"
            f"<td>{(r.get('date') or '')[:10]}</td>"
            f"<td><span style='color:{kind_clr};font-weight:600;'>{kind_lbl}</span></td>"
            f"<td>{r.get('booking_zone','')}</td>"
            f"<td>{r.get('company_name','')}</td>"
            f"<td>{r.get('item_name','')}</td>"
            f"<td>{r.get('time_from','')[:5]}~{r.get('time_to','')[:5]}</td>"
            f"<td>{r.get('vehicle_count','')}대 / {r.get('vehicle_ton','')}t</td>"
            f"<td>{status}</td>"
            f"</tr>"
        )

    generated_at = today_kst().strftime("%Y-%m-%d")

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>자재 반출입 사용현황 보고서</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif; background:#f8fafc; color:#1e293b; }}
  .page {{ max-width:960px; margin:0 auto; padding:32px 24px; }}
  .report-header {{ background:linear-gradient(135deg,#1e3a8a,#2563eb); color:#fff; border-radius:12px; padding:28px 32px; margin-bottom:28px; }}
  .report-header h1 {{ font-size:22px; font-weight:700; margin-bottom:6px; }}
  .report-header p {{ font-size:13px; opacity:0.85; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; margin-bottom:28px; }}
  .kpi-card {{ background:#fff; border-radius:10px; padding:18px 16px; box-shadow:0 1px 4px rgba(0,0,0,0.08); text-align:center; border-top:3px solid #2563eb; }}
  .kpi-card.in  {{ border-top-color:#3b82f6; }}
  .kpi-card.out {{ border-top-color:#f97316; }}
  .kpi-card.done{{ border-top-color:#10b981; }}
  .kpi-num {{ font-size:28px; font-weight:700; color:#1e3a8a; }}
  .kpi-label {{ font-size:12px; color:#64748b; margin-top:4px; }}
  .section {{ background:#fff; border-radius:10px; padding:20px; margin-bottom:20px; box-shadow:0 1px 4px rgba(0,0,0,0.07); }}
  .section-title {{ font-size:14px; font-weight:700; color:#1e3a8a; border-bottom:2px solid #e2e8f0; padding-bottom:10px; margin-bottom:16px; }}
  .bar-chart {{ display:flex; align-items:flex-end; gap:4px; overflow-x:auto; padding-bottom:4px; }}
  table.detail {{ width:100%; border-collapse:collapse; font-size:11px; }}
  table.detail th {{ background:#f1f5f9; color:#475569; padding:7px 8px; text-align:left; font-weight:600; border-bottom:2px solid #e2e8f0; }}
  table.detail td {{ padding:6px 8px; border-bottom:1px solid #f1f5f9; vertical-align:middle; }}
  table.detail tr:hover td {{ background:#f8fafc; }}
  .legend {{ display:flex; gap:16px; font-size:11px; color:#64748b; margin-bottom:10px; }}
  .legend span {{ display:flex; align-items:center; gap:4px; }}
  .dot {{ width:10px; height:10px; border-radius:2px; display:inline-block; }}
  .footer {{ text-align:center; font-size:11px; color:#94a3b8; margin-top:28px; padding-top:16px; border-top:1px solid #e2e8f0; }}
  @media print {{ body{{ background:#fff; }} .page{{ padding:16px; }} }}
</style>
</head>
<body>
<div class="page">

  <div class="report-header">
    <h1>📊 자재 반출입 사용현황 보고서</h1>
    <p>현장 : {site_name} &nbsp;|&nbsp; 기간 : {date_from} ~ {date_to} &nbsp;|&nbsp; 생성일 : {generated_at}</p>
  </div>

  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-num">{total}</div>
      <div class="kpi-label">전체 건수</div>
    </div>
    <div class="kpi-card in">
      <div class="kpi-num" style="color:#3b82f6;">{len(in_reqs)}</div>
      <div class="kpi-label">반입</div>
    </div>
    <div class="kpi-card out">
      <div class="kpi-num" style="color:#f97316;">{len(out_reqs)}</div>
      <div class="kpi-label">반출</div>
    </div>
    <div class="kpi-card done">
      <div class="kpi-num" style="color:#10b981;">{len(done_reqs)}</div>
      <div class="kpi-label">등록완료</div>
    </div>
  </div>

  <div class="section">
    <div class="section-title">📅 일별 반입·반출 추이</div>
    <div class="legend">
      <span><span class="dot" style="background:#3b82f6;"></span>반입</span>
      <span><span class="dot" style="background:#f97316;"></span>반출</span>
    </div>
    <div class="bar-chart">{bar_rows}</div>
  </div>

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px;">
    <div class="section">
      <div class="section-title">🗺 존별 이용 현황</div>
      <table style="width:100%;border-collapse:collapse;">
        {zone_rows}
      </table>
    </div>
    <div class="section">
      <div class="section-title">⏰ 시간대별 집중도</div>
      <table style="width:100%;border-collapse:collapse;">
        {hour_rows}
      </table>
    </div>
  </div>

  <div class="section">
    <div class="section-title">🏢 협력사별 이용 현황 (상위 15)</div>
    <table style="width:100%;border-collapse:collapse;">
      {co_rows}
    </table>
  </div>

  <div class="section">
    <div class="section-title">📋 상세 내역</div>
    <table class="detail">
      <thead>
        <tr>
          <th>날짜</th><th>구분</th><th>존</th><th>협력사</th><th>자재명</th>
          <th>시간</th><th>차량</th><th>상태</th>
        </tr>
      </thead>
      <tbody>{detail_rows}</tbody>
    </table>
  </div>

  <div class="footer">송도역세권아파트 2BL · 자재 반출입 관리 시스템 · {generated_at} 생성</div>
</div>
</body>
</html>"""


def _req_list_for_date(con: Client, project_id: str, target_date: str):
    res = (con.table("requests").select("*")
           .eq("project_id", project_id).eq("date", target_date)
           .order("booking_zone").order("time_from").order("created_at")
           .execute())
    return res.data or []


@measure("page.dashboard")


def page_dashboard(con: Client):
    st.markdown(_DASH_CSS, unsafe_allow_html=True)

    project_id = st.session_state.get("PROJECT_ID", "")
    site_name  = settings_get(con, "site_name", "현장명")

    # ── 날짜 상태 ─────────────────────────────────────────────────────────
    # "dash_date" = 화살표/오늘 버튼이 쓰는 backing store (위젯 key와 분리)
    # "dash_date_picker" = st.date_input 위젯 key
    # Streamlit 규칙: 위젯 key를 외부에서 직접 수정하면 StreamlitAPIException 발생.
    # 해결: backing store로 날짜를 관리하고, date_input은 value= 로만 받음.
    #       캘린더 선택 → on_change 콜백에서 backing store 동기화.
    today: date = today_kst()
    if "dash_date" not in st.session_state:
        st.session_state["dash_date"] = today
    cur_date: date = st.session_state["dash_date"]

    def _sync_picker():
        st.session_state["dash_date"] = st.session_state["dash_date_picker"]

    # ── 날짜 네비게이션 ───────────────────────────────────────────────────
    with st.container(key="dash_nav"):
        nc1, nc2, nc3, nc4, nc5, nc6 = st.columns([1, 1, 2.6, 1, 1, 1])
        with nc1:
            if st.button("◀◀", key="dash_prev_week", use_container_width=True, help="일주일 전"):
                st.session_state["dash_date"] = cur_date - timedelta(days=7)
                st.rerun()
        with nc2:
            if st.button("◀", key="dash_prev_day", use_container_width=True, help="전날"):
                st.session_state["dash_date"] = cur_date - timedelta(days=1)
                st.rerun()
        with nc3:
            st.date_input(
                "날짜", value=cur_date, key="dash_date_picker",
                label_visibility="collapsed",
                on_change=_sync_picker,
            )
        with nc4:
            if st.button("▶", key="dash_next_day", use_container_width=True, help="다음날"):
                st.session_state["dash_date"] = cur_date + timedelta(days=1)
                st.rerun()
        with nc5:
            if st.button("▶▶", key="dash_next_week", use_container_width=True, help="일주일 후"):
                st.session_state["dash_date"] = cur_date + timedelta(days=7)
                st.rerun()
        with nc6:
            is_today = (cur_date == today)
            if st.button(
                "오늘",
                key="dash_today" if not is_today else "dash_today_disabled",
                use_container_width=True,
                type="primary" if not is_today else "secondary",
                disabled=is_today,
                help="이미 오늘입니다" if is_today else "오늘로 이동",
            ):
                st.session_state["dash_date"] = today
                st.rerun()

    # ── 데이터 로드 ───────────────────────────────────────────────────────
    target_str = str(cur_date)
    reqs = _req_list_for_date(con, project_id, target_str)

    # 시간대별(내림차순) → zone별(내림차순) 정렬
    reqs = sorted(reqs, key=lambda r: (r.get("booking_zone", ""), r.get("kind", ""), r.get("time_from", "")))

    dow_map = {0:"월요일", 1:"화요일", 2:"수요일", 3:"목요일",
               4:"금요일", 5:"토요일", 6:"일요일"}
    date_label = f"{cur_date.year}년 {cur_date.month}월 {cur_date.day}일 {dow_map[cur_date.weekday()]}"

    # ── 터미널 사용 존 목록 로드 ─────────────────────────────────────────
    import json as _json
    try:
        _terminal_zones = set(_json.loads(settings_get(con, "terminal_zones_json", '[]')))
    except Exception:
        _terminal_zones = set()

    # ── 테이블 행 생성 ────────────────────────────────────────────────────
    row_parts = []
    total_cnt = 0

    if not reqs:
        row_parts.append('<tr><td colspan="12" style="padding:30px;color:#94a3b8;">해당 날짜에 등록된 요청이 없습니다.</td></tr>')
    else:
        for i, r in enumerate(reqs, 1):
            kind     = r.get("kind", KIND_IN)
            is_in    = kind == KIND_IN
            kind_cls = "kind-in" if is_in else "kind-out"
            kind_lbl = "반입" if is_in else "반출"
            zone     = r.get("booking_zone", "")
            company  = r.get("company_name", "")
            item     = r.get("item_name", "")
            vcnt_raw = r.get("vehicle_count", "")
            vcnt     = f"{vcnt_raw}대" if vcnt_raw else ""
            vton     = r.get("vehicle_ton", "")
            loading  = r.get("loading_method", "")
            gate_raw = r.get("gate", "")
            gate_raw = "" if gate_raw == "선택" else gate_raw
            # 터미널 미사용 존이면 N/A 표기
            gate     = gate_raw if zone in _terminal_zones else "N/A"
            t_from = r.get("time_from", "")
            sup    = r.get("worker_supervisor", "")
            guide  = r.get("worker_guide", "")
            mgr    = r.get("worker_manager", "")
            vcnt_int = int(vcnt_raw) if str(vcnt_raw).isdigit() else 0
            total_cnt += vcnt_int

            tds = (
                f'<td data-content="{i}" class="tooltip-cell">{i}</td>'
                f'<td data-content="{zone}" class="tooltip-cell">{zone}</td>'
                f'<td data-content="{gate}" class="tooltip-cell">{gate}</td>'
                f'<td data-content="{company}" class="tooltip-cell">{company}</td>'
                f'<td data-content="{item}" class="tooltip-cell">{item}</td>'
                f'<td data-content="{vcnt}" class="tooltip-cell">{vcnt}</td>'
                f'<td data-content="{kind_lbl} / {vton}" class="tooltip-cell"><span class="{kind_cls}">{kind_lbl}</span> / {vton}</td>'
                f'<td data-content="{loading}" class="tooltip-cell">{loading}</td>'
                f'<td data-content="{t_from}" class="tooltip-cell">{t_from}</td>'
                f'<td data-content="{sup}" class="tooltip-cell">{sup}</td>'
                f'<td data-content="{guide}" class="tooltip-cell">{guide}</td>'
                f'<td data-content="{mgr}" class="tooltip-cell">{mgr}</td>'
            )
            row_parts.append(f'<tr>{tds}</tr>')

        # 합계 행
        row_parts.append(
            f'<tr class="total-row">'
            f'<td colspan="5" style="text-align:center;padding-right:8px;">합 계</td>'
            f'<td>{total_cnt}대</td>'
            f'<td colspan="6"></td>'
            f'</tr>'
        )

    rows_html = "".join(row_parts)

    # ── HTML 렌더링 ───────────────────────────────────────────────────────
    thead = (
        '<thead>'
        '<tr>'
        '<th>No</th>'
        '<th>Zone</th>'
        '<th>터미널</th>'
        '<th>업체명</th>'
        '<th>자재종류</th>'
        '<th>수량</th>'
        '<th>반입·반출<br>차량</th>'
        '<th>상·하차<br>방식</th>'
        '<th>시간</th>'
        '<th>작업<br>지휘자</th>'
        '<th>유도원</th>'
        '<th>담당자</th>'
        '</tr>'
        '</thead>'
    )
    cnt_summary = f"총 {len(reqs)}건 (반입 {sum(1 for r in reqs if r.get('kind')==KIND_IN)}건 / 반출 {sum(1 for r in reqs if r.get('kind')==KIND_OUT)}건)" if reqs else "등록된 요청 없음"
    mobile_html = (
        f'<div class="dash-mobile-only">'
        f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:20px;text-align:center;">'
        f'<div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:6px;">자재 상·하차 반입/반출 현황</div>'
        f'<div style="font-size:13px;color:#475569;margin-bottom:4px;">□ 현장명 : {site_name}</div>'
        f'<div style="font-size:13px;color:#475569;margin-bottom:12px;">{date_label}</div>'
        f'<div style="font-size:13px;color:#2563eb;font-weight:600;">{cnt_summary}</div>'
        f'</div>'
        f'</div>'
    )
    html = (
        '<div class="dash-wrap">'
        '<div class="dash-title-box"><h2>자재 상 · 하차 반입 / 반출 현황</h2></div>'
        f'<div class="dash-meta"><span>□ 현장명 : {site_name}</span><span>{date_label}</span></div>'
        f'<table class="dash-table">{thead}<tbody>{rows_html}</tbody></table>'
        '</div>'
        f'{mobile_html}'
    )
    st.markdown(html, unsafe_allow_html=True)

    # 드래그-스크롤: .dash-wrap 위에서 마우스 클릭 드래그로 가로 이동
    st.markdown("""
<script>
(function(){
  function initDrag(el){
    let isDown=false, startX=0, scrollLeft=0;
    el.addEventListener('mousedown',function(e){
      if(e.button!==0) return;
      isDown=true; startX=e.pageX-el.offsetLeft; scrollLeft=el.scrollLeft;
      e.preventDefault();
    });
    el.addEventListener('mouseleave',function(){ isDown=false; });
    el.addEventListener('mouseup',function(){ isDown=false; });
    el.addEventListener('mousemove',function(e){
      if(!isDown) return;
      var x=e.pageX-el.offsetLeft, walk=(x-startX)*1.2;
      el.scrollLeft=scrollLeft-walk;
    });
  }
  function tryInit(){
    var el=document.querySelector('.dash-wrap');
    if(el){ initDrag(el); } else { setTimeout(tryInit,300); }
  }
  tryInit();
})();
</script>

<script>
(function(){
  let activeTooltip = null;
  let lastClickedCell = null;
  let lastClickTime = 0;

  document.addEventListener('click', function(e){
    const cell = e.target.closest('.tooltip-cell');

    // 말풍선 클릭 → 사라짐
    if(e.target.classList && e.target.classList.contains('tooltip-content')){
      if(activeTooltip) activeTooltip.remove();
      activeTooltip = null;
      lastClickedCell = null;
      return;
    }

    // 테이블 셀 클릭
    if(cell){
      e.stopPropagation();
      const now = Date.now();
      const isDoubleClick = lastClickedCell === cell && (now - lastClickTime) < 300;

      if(isDoubleClick){
        // 더블클릭 → 페이지 이동 (요청 ID 기반)
        const row = cell.closest('tr');
        const reqData = row.getAttribute('data-req-id');
        if(reqData){
          window.location.href = '?page=approval&rid=' + reqData;
        }
      } else {
        // 싱글클릭 → 말풍선 표시
        const content = cell.getAttribute('data-content') || cell.textContent;
        if(activeTooltip) activeTooltip.remove();

        const tooltip = document.createElement('div');
        tooltip.className = 'tooltip-content active';
        tooltip.textContent = content;

        const rect = cell.getBoundingClientRect();
        tooltip.style.top = (rect.top - 60) + 'px';
        tooltip.style.left = (rect.left + rect.width/2 - 140) + 'px';

        document.body.appendChild(tooltip);
        activeTooltip = tooltip;
      }

      lastClickedCell = cell;
      lastClickTime = now;
    }
  });
})();
</script>
""", unsafe_allow_html=True)

    # ── 엑셀 다운로드 버튼 ────────────────────────────────────────────────
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    if reqs:
        try:
            excel_bytes = _build_excel(reqs, site_name, date_label, _terminal_zones)
            filename = f"반입반출현황_{target_str}.xlsx"
            _, btn_col, _ = st.columns([1.5, 2, 1.5])
            with btn_col:
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        except ImportError:
            st.warning("엑셀 다운로드를 사용하려면 `pip install openpyxl` 을 실행하세요.")

    # ── 사용현황 보고서 다운로드 ─────────────────────────────────────────────
    st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
    with st.expander("📊 기간별 사용현황 보고서"):
        r_col1, r_col2 = st.columns(2)
        with r_col1:
            from datetime import date as _d
            r_from = st.date_input("시작일", value=_d(2026, 6, 15), key="report_from")
        with r_col2:
            r_to = st.date_input("종료일", value=_d.today(), key="report_to")
        if st.button("📊 보고서 생성 및 다운로드", key="report_gen", use_container_width=True, type="primary"):
            r_from_str = str(r_from)
            r_to_str   = str(r_to)
            _res = (con.table("requests").select("*")
                    .eq("project_id", project_id)
                    .gte("date", r_from_str)
                    .lte("date", r_to_str)
                    .order("date").order("booking_zone").order("time_from")
                    .execute())
            _all = _res.data or []
            if _all:
                _html = _build_report_html(_all, site_name, r_from_str, r_to_str)
                st.download_button(
                    label="⬇️ HTML 파일 저장",
                    data=_html.encode("utf-8"),
                    file_name=f"사용현황보고서_{r_from_str}~{r_to_str}.html",
                    mime="text/html",
                    use_container_width=True,
                )
            else:
                st.warning("해당 기간에 데이터가 없습니다.")
